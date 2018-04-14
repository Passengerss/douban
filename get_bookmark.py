''' 将标签信息储存到表格 '''
import urllib.request
import re
from lxml import etree      # 使用xpath筛选器
from user_agent.base import generate_user_agent
from openpyxl import Workbook, load_workbook      # 用于创建 和 读取 表格文件
from openpyxl.styles import colors, Font, Alignment, Border, Side       # 改变字体颜色，大小, 对其方式, 边框


def url_open(url):
    head = {"User-Agent": generate_user_agent()}
    req = urllib.request.Request(url, headers=head)
    response = urllib.request.urlopen(req).read()
    return response


# 获取首页书签，并存为表格
def get_mark(url):
    response = url_open(url).decode('utf-8')
    html = etree.HTML(response, parser=None, )
    # 获取标签总分类的列表
    categories = html.xpath('//a[@class="tag-title-wrapper"]/@name')

    # 获取《文学》下的标签分类
        # 匹配出文学分类下包含的所有标签内容
    literature_string = re.compile('(<a name="文学" class="tag-title-wrapper">\s(\s|.)*?\s</div>)').findall(response)[0][0]
    # 实例化对象为可xpath操作的对象，xpath返回列表
    html = etree.HTML(literature_string, parser=None, )
    literature_string = html.xpath('//td/a/text()')
    # 保存第一个标签分类
    yield save_main_mark(row=main_row_start, value=categories[0])
    # 保存第一个标签下的子标签
    yield save_mark(literature_string)

    # 获取《流行》下的标签分类
    popular_string = re.compile('(<a name="流行" class="tag-title-wrapper">\s(\s|.)*?\s</div>)').findall(response)[0][0]
    html = etree.HTML(popular_string, parser=None, )
    popular_string = html.xpath('//td/a/text()')
    yield save_main_mark(row=main_row_start, value=categories[1])
    yield save_mark(popular_string)

    # 获取《文化》下的标签分类
    culture_string = re.compile('(<a name="文化" class="tag-title-wrapper">\s(\s|.)*?\s</div>)').findall(response)[0][0]
    html = etree.HTML(culture_string, parser=None, )
    culture_string = html.xpath('//td/a/text()')
    yield save_main_mark(row=main_row_start, value=categories[2])
    yield save_mark(culture_string)

    # 获取《生活》下的标签分类
    life_string = re.compile('(<a name="生活" class="tag-title-wrapper">\s(\s|.)*?\s</div>)').findall(response)[0][0]
    html = etree.HTML(life_string, parser=None, )
    life_string = html.xpath('//td/a/text()')
    yield save_main_mark(row=main_row_start, value=categories[3])
    yield save_mark(life_string)

    # 获取《经管》下的标签分类
    manage_string = re.compile('(<a name="经管" class="tag-title-wrapper">\s(\s|.)*?\s</div>)').findall(response)[0][0]
    html = etree.HTML(manage_string, parser=None, )
    manage_string = html.xpath('//td/a/text()')
    yield save_main_mark(row=main_row_start, value=categories[4])
    yield save_mark(manage_string)

    # 获取《科技》下的标签分类
    technology_string = re.compile('(<a name="科技" class="tag-title-wrapper">\s(\s|.)*?\s</div>)').findall(response)[0][0]
    html = etree.HTML(technology_string, parser=None, )
    technology_string = html.xpath('//td/a/text()')
    yield save_main_mark(row=main_row_start, value=categories[5])
    yield save_mark(technology_string)


# 传入分类value,储存主分类，文学，流行...etc
def save_main_mark(row, value):
    # 如果是第一次就创建表格
    if main_row_start == 1:
        wb = Workbook()     # 创建工作表
        ws1 = wb.active  # x选中工作表中的第一个sheet，_active_sheet_index属性默认为 0
        ws1.title = "标签"  # 更改sheet1的名字为标签
        ws1.sheet_properties.tabColor = "1072BA"  # sheet1的背景颜色，有RRGGBB确定
    else:
        wb = load_workbook(filename="豆瓣图书.xlsx",)
    ws1 = wb.active     # x选中工作表中的第一个sheet，_active_sheet_index属性默认为 0

    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)    # 合并单元格 A1-D1 ,ws1.merge_cells('A1:D4')
    ws1.cell(column=1, row=row, value=value).font=Font(color=colors.RED, italic=None, size=18, bold=True,)      # 操作单元格,红色，加粗
    ws1.cell(column=1, row=row, value=value).alignment = Alignment(horizontal="center", vertical="center")
    wb.save(filename="豆瓣图书.xlsx")


row_start = 2     # 从第二行开始
main_row_start = 1  # 标签分类从一行开始

# 存储子分类。
def save_mark(x):
    # 定义全局变量方便每次都能按顺序自动存储
    global row_start
    global main_row_start
    tag = 0
    wb = load_workbook(filename="豆瓣图书.xlsx",)
    ws1 = wb.active
    # 设置第row行行高为30
    # ws1.row_dimensions[row].height = 30
    # 设置第col列列宽为20
    for col in 'ABCD':
        ws1.column_dimensions[col].width = 20

    # 储存为 7*4 的表格
    for row in range(row_start, row_start+10):
        for col in range(1, 5):
            # 设置字体
            ws1.cell(row=row, column=col, value=x[tag]).font = Font(color='EE6A50', size=14, bold=True,)
            # 设置对其格式
            ws1.cell(row=row, column=col, value=x[tag]).alignment = Alignment(horizontal="center", vertical="center")
            # 设置边界样式
            ws1.cell(row=row, column=col, value=x[tag]).border = Border(
                top=Side(color='EE6A50'), left=Side(color='EE6A50'), right=Side(color='EE6A50'), bottom=Side(color='EE6A50'))

            tag += 1        # 将标签内容移位
            if tag == len(x):       # 判断数据是否保存完
                main_row_start = row + 1    #
                row_start = main_row_start + 1  #
                wb.save(filename="豆瓣图书.xlsx")
                return None


if __name__ == "__main__":
    url = "https://book.douban.com/tag/"

    save = get_mark(url)
    print("正在处理，请稍后...")
    for fun in save:    # for循环会自动调用 next()方法，和处理StopIteration(溢出)异常
        pass
    print("--------------Finished!---------------")


