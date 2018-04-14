'''
格式为：title，author，rating_nums, comment_nums, link
'''
import urllib.request, urllib.parse
import urllib
from user_agent.base import generate_user_agent
import threading
from lxml import etree
from openpyxl import Workbook, load_workbook    # 导入创建和加载工作簿的库文件
from openpyxl.styles import Font, colors, Alignment     # 导入工作簿样式所需的库文件，字体，颜色，对齐方式
from openpyxl.worksheet.table import Table, TableStyleInfo      # 导入工作表中制作表格所需的库文件，table， 表格风格
import queue        # 产生队列，先进先出.Queue()
information = queue.Queue()
url_queue = queue.Queue()
import time
import re

# 获取内容标签下某一页内容
class Book(threading.Thread):
    def __init__(self, url, queue):
        super().__init__()
        self.url = url
        self.queue = queue

    def url_open(self, url):
        head = {"User-Agent" : generate_user_agent()}
        req = urllib.request.Request(url, headers=head)
        response = urllib.request.urlopen(req).read()
        return response

    def run(self):
        print("正在获取网站信息...\n地址：%s\n" % self.url)
        response = self.url_open(self.url).decode('utf-8')
        # 使用HTML解析网页
        response = etree.HTML(response, parser=None)

        book_name = response.xpath('//h2/a/@title')
        # 对获取的数据进行格式处理
        book_author = []
        book_authors = response.xpath('//div[@class="pub"]/text()')
        for each in book_authors:
            temp = each.replace(' ','').strip()
            book_author.append(temp.split("/")[0])

        rating_num = []
        rating_nums = response.xpath('//span[@class="rating_nums"]/text()')
        for each in rating_nums:
            rating_num.append(float(each))

        comment_num = []
        comment_nums = response.xpath('//span[@class="pl"]/text()')
        for each in comment_nums:
            comment_num.append(int(''.join(re.findall(r'[0-9]', each))))

        book_link = response.xpath('//h2/a/@href')

        info_zip = list(zip(book_name, book_author, rating_num, comment_num, book_link))
        self.queue.put(info_zip,timeout=None,)
        self.queue.task_done()

# 用来保存某一页的内容
class Save(threading.Thread):
    def __init__(self, queue, filename, sheetname):
        super().__init__()
        self.queue = queue
        self.filename = filename
        self.sheetname = sheetname

    def run(self):
        i = 2
        wb = load_workbook(filename=self.filename)  # 打开已有表格
        ws2 = wb.create_sheet(title=self.sheetname)
        # 添加第一行数据
        ws2.merge_cells("E1:I1")
        ws2.append(("书名", "作者", "豆瓣评分", "评价人数", "链接地址"))

        # 判断数据是否存完，True则等待，False 则继续
        while self.queue.empty() is False:
            print("正在保存数据到表格，请稍后...")
            content = self.queue.get()      # 从队列中取数据
            print("此页包含 %d 个数据" %len(content))
            for row in content:
                ws2.merge_cells(start_column=5, start_row=i, end_column=9, end_row=i)
                i += 1
                ws2.append(row) # 添加行数据
        wb.save(filename=self.filename)
        print("数据保存完成！")
        hanlder.start()
        hanlder.join()

# 用来处理文档的格式
class Hanlder(threading.Thread):
    def __init__(self, queue, filename, sheetname):
        super().__init__()
        self.queue = queue
        self.filename = filename
        self.sheetname = sheetname

    def run(self):
        try:
            if self.queue.empty() is True:     # 若数据已保存完毕
                wb = load_workbook(filename=self.filename)  # 打开已有表格
                print("正在处理表格样式，请稍后...")
                ws2 = wb[self.sheetname]     # 选择某一个工作表 ws2 = wb.worksheets[1] 通过索引获取或名字wb["name"]
                ws2.sheet_properties.tabColor = "DDA0DD"        # 设置工作表背景色
                # 设置列宽
                ws2.column_dimensions['A'].width = 20
                ws2.column_dimensions['B'].width = 28
                ws2.column_dimensions['C'].width = 12
                ws2.column_dimensions['D'].width = 15
                ws2.column_dimensions["E"].width = 20
                # 设置对齐与字体
                for cells in ws2["A1:E1"]:
                    for cell in cells:
                        cell.alignment = Alignment(horizontal="center", vertical="center") # 第一行对齐
                        cell.font = Font(size=16, color="A020F0", bold=True)

                row_length = (ws2.max_row)  # 获得包含数据的总行数int sheet.rows 为返回所有行(含数据)可用来迭代
                col_length = ws2.max_column # 获取总列数
                for row in range(2, row_length+1):
                    for col in range(1, col_length+1):
                        ws2.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center") # 居中对齐
                        ws2.cell(row=row, column=col).font = Font(size=12, color="B452CD", bold=True)   # 字体
                    ws2.cell(row=row, column=col_length).alignment = Alignment(horizontal="left", vertical="center")  # 最后一列左对齐
                '''
                 因为 表中包含数字，而openpyxl规定，ref过滤器范围必须始终包含字符串，否则Excel会报错并删除表格,应该是这样
                # 制作成表格
                # Add a default style with striped rows and banded columns
                style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                tab = Table(displayName="书名信息表", ref="A1:C21",tableStyleInfo=style)
                ws2.add_table(tab)
                '''
                wb.save(self.filename)
                print("格式处理完成！")
        except PermissionError as e:
            print("表格已被打开，请先关闭! %s" %e)

if __name__ == "__main__":
    tag = "小说"
    tag_encode = urllib.parse.quote(tag)   # 编码中文
    filename = "豆瓣图书.xlsx"
    sheetname = tag
    for page in range(0, 2000, 20):
        url = "https://book.douban.com/tag/{tag}?start={int}".format(tag=tag_encode, int=page)
        Book(url=url, queue=information).start()
    save = Save(queue=information, filename=filename, sheetname=sheetname)
    hanlder = Hanlder(queue=information, filename=filename, sheetname=sheetname)
    time.sleep(1)
    save.start()

